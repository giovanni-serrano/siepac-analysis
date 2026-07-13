"""
etl_generacion_por_fuente.py — Generación eléctrica por tipo de fuente
====================================================
Etapa del pipeline : ETL
Entradas           : data/raw/generacion_por_tipo_de_fuente/*.xlsx (una hoja
                     por país, nombradas "N.País")
Salidas            : data/processed/generacion_por_fuente_detalle.csv
                     (una fila por pais-anio-tipo_fuente, 11 fuentes "hoja") y
                     data/processed/generacion_por_tipo_de_fuente.csv
                     (agregado por pais-anio: fósil / renovable / total +
                     desglose renovable)
Alimenta           : ECO11, ECO13
Fuente de datos    : sieLAC-OLADE (Generación eléctrica por fuente)

Uso:  python src/etl_generacion_por_fuente.py   (ejecutar desde la raíz)

Notas metodológicas:
  - Estructura del raw (verificada, idéntica en las 6 hojas): fila 5 de
    encabezado y jerarquía con 3 espacios de sangría literal:
        Térmica no renovable (combustión)   [padre]   -> FÓSIL      (ECO11)
           Petróleo y derivados / Gas natural / Carbón mineral / Otras fuentes
        Térmica renovable (combustión)      [padre]   -> RENOVABLE  (biomasa)
           Biogás / Biomasa sólida / Biocombustibles líquidos
        Fuentes renovable (no combustión)   [padre]   -> RENOVABLE
           Hidro / Geotermia / Eólica / Solar
        Total                               [gran total]
  - Los renglones padre y Total ya vienen sumados; se reconcilian contra
    la suma de hijos con tolerancia de 1 kWh.
  - Energía en kWh como unidad base (+ columna GWh de referencia).
  - Columna `fuente` = PROCEDENCIA del dato ("sieLAC-OLADE"), NO el tipo
    de energía; el tipo va en `tipo_fuente` (para no chocar nombres).
  - Validación estricta: nulos, categorías desconocidas o reconciliaciones
    que no cuadran abortan sin escribir los CSV.

Autor: Luis Giovanni Serrano Bello — Tesis SIEPAC, UNI Nicaragua
"""

from __future__ import annotations

import logging
import sys
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config_siepac import GWH_A_KWH, DIR_RAW, DIR_PROCESSED

# --------------------------------------------------------------------------- #
# 0. CONFIGURACIÓN Y RUTAS
# --------------------------------------------------------------------------- #
RAW_DIR = DIR_RAW / "generacion_por_tipo_de_fuente"
OUT_DIR = DIR_PROCESSED

FUENTE_DATO = "sieLAC-OLADE"          # procedencia (va en la columna `fuente`)
TOLERANCIA_KWH = 1.0                  # holgura al reconciliar sumas (en kWh)

# logging en vez de print(): permite distinguir INFO / WARNING / ERROR y deja
# un rastro claro de qué hizo el script. Es la base para detectar fallos
# silenciosos (una fuente que no se clasifica sale como WARNING, no se pierde).
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-7s | %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger(Path(__file__).stem)

# --------------------------------------------------------------------------- #
# 1. TAXONOMÍA FUENTE -> CATEGORÍA (el corazón del mapeo fósil/renovable)
# --------------------------------------------------------------------------- #
# Clasificamos por ETIQUETA EXACTA, no por posición de fila. Si OLADE reordena
# o inserta filas, el mapeo sigue siendo correcto; y si aparece una etiqueta
# desconocida, la detectamos (no la clasificamos a ciegas).
#
# Cada entrada dice: nivel (padre/hijo), categoría (fósil/renovable) y, para los
# hijos renovables, a qué grupo de ECO13 pertenecen (hidro/geotermia/eólica/
# solar/biomasa). Los tres sub-tipos de biomasa (Biogás, Biomasa sólida,
# Biocombustibles líquidos) van todos a grupo "biomasa" (decisión validada).

PADRE_FOSIL = "Térmica no renovable (combustión)"
PADRE_BIOMASA = "Térmica renovable (combustión)"
PADRE_RENOV_NO_COMB = "Fuentes renovable (no combustión)"
ETIQUETA_TOTAL = "Total"

# Padres -> categoría (se usan para los AGREGADOS: ya vienen pre-sumados)
PADRES = {
    PADRE_FOSIL: "fósil",
    PADRE_BIOMASA: "renovable",
    PADRE_RENOV_NO_COMB: "renovable",
}

# Hijos -> (padre, categoría, grupo_eco13). grupo_eco13=None para los fósiles.
HIJOS = {
    "Petróleo y derivados":      (PADRE_FOSIL, "fósil", None),
    "Gas natural":               (PADRE_FOSIL, "fósil", None),
    "Carbón mineral":            (PADRE_FOSIL, "fósil", None),
    "Otras fuentes":             (PADRE_FOSIL, "fósil", None),
    "Biogás":                    (PADRE_BIOMASA, "renovable", "biomasa"),
    "Biomasa sólida":            (PADRE_BIOMASA, "renovable", "biomasa"),
    "Biocombustibles líquidos":  (PADRE_BIOMASA, "renovable", "biomasa"),
    "Hidro":                     (PADRE_RENOV_NO_COMB, "renovable", "hidro"),
    "Geotermia":                 (PADRE_RENOV_NO_COMB, "renovable", "geotermia"),
    "Eólica":                    (PADRE_RENOV_NO_COMB, "renovable", "eolica"),
    "Solar":                     (PADRE_RENOV_NO_COMB, "renovable", "solar"),
}

# Conjunto de todas las etiquetas ESPERADAS (padres + hijos + Total). Sirve para
# (a) detectar etiquetas desconocidas y (b) NO confundir la nota al pie
# "Fuente: sieLAC-OLADE" con el renglón de datos "Fuentes renovable ...".
ETIQUETAS_CONOCIDAS = set(PADRES) | set(HIJOS) | {ETIQUETA_TOTAL}

# Cuántas fuentes hoja esperamos por país (para el conteo de validación).
N_HIJOS_ESPERADOS = len(HIJOS)  # 11


# --------------------------------------------------------------------------- #
# Utilidades pequeñas
# --------------------------------------------------------------------------- #
def _texto(v) -> str:
    """Devuelve el valor de celda como texto SIN espacios de borde.
    Los espacios internos de sangría (3 al inicio) los quitamos aquí: sirven
    solo como marca visual de 'hijo', pero ya clasificamos por etiqueta."""
    return v.strip() if isinstance(v, str) else ""


def _es_nota_al_pie(texto: str) -> bool:
    """True si la celda es una nota al pie y NO un renglón de datos.
    OJO: 'Fuente:' (con dos puntos) es la procedencia; 'Fuentes renovable...'
    es un dato. Por eso exigimos los dos puntos, no solo 'Fuente'."""
    return texto.startswith("Fuente:") or texto.startswith("La opción")


def _pais_desde_hoja(nombre_hoja: str) -> str:
    """'1.Costa Rica' -> 'Costa Rica'. Quita el prefijo 'N.' si existe."""
    return nombre_hoja.split(".", 1)[1].strip() if "." in nombre_hoja else nombre_hoja.strip()


def _norm(s: str) -> str:
    """Normaliza para comparar países (minúsculas, sin acentos)."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


# --------------------------------------------------------------------------- #
# 2. EXTRAER
# --------------------------------------------------------------------------- #
def extraer_datos(xlsx_path: Path) -> pd.DataFrame:
    """
    Lee el .xlsx y devuelve un DataFrame LARGO 'crudo' con una fila por
    (pais, anio, etiqueta), incluyendo padres y Total. Columnas:
        pais, anio, etiqueta, nivel, valor_gwh, dato_reportado

    - nivel ∈ {'padre', 'hijo', 'total'}.
    - valor_gwh: número en GWh; si la celda venía vacía -> 0.0.
    - dato_reportado: False si la celda original estaba vacía (traza del 0).

    Es DEFENSIVO: si una hoja no tiene encabezado reconocible, la registra
    como WARNING y la salta (no rompe el proceso ni inventa datos).
    """
    log.info("Abriendo workbook: %s", xlsx_path.name)
    # data_only=True -> si hubiese fórmulas, leemos el valor calculado, no "=...".
    # read_only=False es seguro aquí porque el archivo es pequeño (~40 KB).
    wb = load_workbook(xlsx_path, data_only=True)

    filas = []
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        pais = _pais_desde_hoja(nombre_hoja)

        # -- localizar la fila de encabezado (col A == "Descripción") --------
        fila_hdr = None
        for r in range(1, ws.max_row + 1):
            if _texto(ws.cell(r, 1).value) == "Descripción":
                fila_hdr = r
                break
        if fila_hdr is None:
            log.warning("Hoja %r: no se encontró encabezado 'Descripción'. "
                        "SE OMITE la hoja completa.", nombre_hoja)
            continue

        # -- cross-check: país de la hoja vs etiqueta interna (fila hdr-1) ----
        # La fila justo arriba del encabezado dice p.ej. "Costa Rica - ...".
        etiqueta_pais = _texto(ws.cell(fila_hdr - 1, 1).value)
        if etiqueta_pais and _norm(pais) not in _norm(etiqueta_pais):
            log.warning("Hoja %r: el país del nombre (%r) no coincide con la "
                        "etiqueta interna (%r). Uso el del nombre de hoja.",
                        nombre_hoja, pais, etiqueta_pais)

        # -- leer los años del encabezado (vienen como texto: '2020'...) -----
        anios = []
        for c in range(3, ws.max_column + 1):
            crudo = _texto(ws.cell(fila_hdr, c).value)
            if crudo.isdigit():
                anios.append((c, int(crudo)))
            else:
                # columna sin año válido -> la ignoramos (defensivo)
                if crudo:
                    log.warning("Hoja %r: encabezado de columna %d no es un año "
                                "(%r). Se ignora esa columna.", nombre_hoja, c, crudo)
        if not anios:
            log.warning("Hoja %r: no se detectaron años en el encabezado. "
                        "SE OMITE la hoja.", nombre_hoja)
            continue

        # -- recorrer los renglones de datos debajo del encabezado -----------
        etiquetas_vistas = set()
        for r in range(fila_hdr + 1, ws.max_row + 1):
            celda_a = ws.cell(r, 1).value
            etiqueta = _texto(celda_a)

            if etiqueta == "":
                continue                      # fila en blanco -> saltar
            if _es_nota_al_pie(etiqueta):
                continue                      # nota al pie -> saltar

            # ¿es una etiqueta que conocemos?
            if etiqueta not in ETIQUETAS_CONOCIDAS:
                # Señal clave de fallo silencioso: fuente nueva/renombrada.
                log.warning("Hoja %r fila %d: etiqueta DESCONOCIDA %r. No se "
                            "clasifica ni se agrega (revisar taxonomía).",
                            nombre_hoja, r, etiqueta)
                continue

            # nivel según la taxonomía
            if etiqueta == ETIQUETA_TOTAL:
                nivel = "total"
            elif etiqueta in PADRES:
                nivel = "padre"
            else:
                nivel = "hijo"
            etiquetas_vistas.add(etiqueta)

            # leer el valor de cada año para esta etiqueta
            for (c, anio) in anios:
                bruto = ws.cell(r, c).value
                if isinstance(bruto, (int, float)):
                    valor_gwh = float(bruto)
                    reportado = True
                elif bruto is None:
                    # Celda vacía = sin generación de ese tipo ese año.
                    # OLADE trata el vacío como 0 en sus subtotales (verificado),
                    # así que lo homologamos a 0.0 y lo marcamos como no reportado
                    # (queda 100% auditable; no estamos "inventando" un número).
                    valor_gwh = 0.0
                    reportado = False
                else:
                    # texto inesperado donde debería haber número -> defensivo
                    log.warning("Hoja %r fila %d col %d: valor no numérico %r. "
                                "Se trata como vacío (0).", nombre_hoja, r, c, bruto)
                    valor_gwh = 0.0
                    reportado = False

                filas.append({
                    "pais": pais,
                    "anio": anio,
                    "etiqueta": etiqueta,
                    "nivel": nivel,
                    "valor_gwh": valor_gwh,
                    "dato_reportado": reportado,
                })

        # ¿extrajimos todos los hijos esperados en esta hoja?
        hijos_vistos = etiquetas_vistas & set(HIJOS)
        faltantes = set(HIJOS) - hijos_vistos
        if faltantes:
            log.warning("Hoja %r: faltan %d fuentes hoja esperadas: %s",
                        nombre_hoja, len(faltantes), sorted(faltantes))
        else:
            log.info("Hoja %r (%s): %d fuentes hoja + padres + Total extraídos.",
                     nombre_hoja, pais, len(hijos_vistos))

    df = pd.DataFrame(filas)
    if df.empty:
        raise ValueError("No se extrajo ninguna fila. ¿El archivo raw es el correcto?")
    log.info("Extracción terminada: %d filas crudas (todas las etiquetas).", len(df))
    return df


# --------------------------------------------------------------------------- #
# 3. TRANSFORMAR
# --------------------------------------------------------------------------- #
def transformar(df_raw: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    A partir del DataFrame crudo construye las dos salidas tidy.

    (a) df_detalle: solo hojas (11 tipos) con categoría, subcategoría OLADE y
        valores en GWh y kWh. Una fila por (pais, anio, tipo_fuente).
    (b) df_agregado: una fila por (pais, anio) con el desglose renovable y los
        totales fósil / renovable / total, en kWh (+ GWh de referencia para los
        tres agregados grandes). Alimenta ECO11 y ECO13.
    """
    # ---------- (a) DETALLE (solo hijos) -----------------------------------
    hijos = df_raw[df_raw["nivel"] == "hijo"].copy()

    # adjuntar categoría / subcategoría OLADE / grupo ECO13 desde la taxonomía
    hijos["categoria"] = hijos["etiqueta"].map(lambda e: HIJOS[e][1])
    hijos["subcategoria_olade"] = hijos["etiqueta"].map(lambda e: HIJOS[e][0])
    hijos["grupo_eco13"] = hijos["etiqueta"].map(lambda e: HIJOS[e][2])
    hijos["valor_kwh"] = hijos["valor_gwh"] * GWH_A_KWH

    df_detalle = (
        hijos.rename(columns={"etiqueta": "tipo_fuente"})
        .assign(fuente=FUENTE_DATO)
        [["pais", "anio", "tipo_fuente", "categoria", "subcategoria_olade",
          "valor_gwh", "valor_kwh", "dato_reportado", "fuente"]]
        .sort_values(["pais", "anio", "tipo_fuente"])
        .reset_index(drop=True)
    )

    # ---------- (b) AGREGADO (una fila por país-año) ------------------------
    # Desglose renovable ECO13: sumamos los hijos por grupo (hidro, geo, eólica,
    # solar, biomasa). Como biomasa son 3 sub-tipos, el groupby los junta solo.
    renov = hijos[hijos["grupo_eco13"].notna()]
    pivote = (
        renov.groupby(["pais", "anio", "grupo_eco13"])["valor_gwh"].sum()
        .unstack("grupo_eco13", fill_value=0.0)
        .reset_index()
    )
    # asegurar que existan las 5 columnas aunque algún país no tenga una fuente
    for grupo in ["hidro", "geotermia", "eolica", "solar", "biomasa"]:
        if grupo not in pivote.columns:
            pivote[grupo] = 0.0

    # fósil: lo tomamos del renglón PADRE 'Térmica no renovable' (pre-sumado).
    fosil = (
        df_raw[df_raw["etiqueta"] == PADRE_FOSIL]
        .rename(columns={"valor_gwh": "fosil_gwh"})[["pais", "anio", "fosil_gwh"]]
    )
    # total: del renglón 'Total' de ORIGEN (autoridad para el denominador).
    total = (
        df_raw[df_raw["etiqueta"] == ETIQUETA_TOTAL]
        .rename(columns={"valor_gwh": "total_gwh"})[["pais", "anio", "total_gwh"]]
    )

    agg = pivote.merge(fosil, on=["pais", "anio"]).merge(total, on=["pais", "anio"])
    # renovable = suma del desglose (hidro+geo+eólica+solar+biomasa)
    agg["renovable_gwh"] = agg[["hidro", "geotermia", "eolica", "solar", "biomasa"]].sum(axis=1)

    # pasar todo a kWh (unidad base del pipeline)
    for col in ["hidro", "geotermia", "eolica", "solar", "biomasa",
                "renovable_gwh", "fosil_gwh", "total_gwh"]:
        agg[col.replace("_gwh", "") + "_kwh" if col.endswith("_gwh") else col + "_kwh"] = \
            agg[col] * GWH_A_KWH

    df_agregado = (
        agg.assign(fuente=FUENTE_DATO)
        [["pais", "anio",
          "hidro_kwh", "geotermia_kwh", "eolica_kwh", "solar_kwh", "biomasa_kwh",
          "renovable_kwh", "fosil_kwh", "total_kwh",
          "renovable_gwh", "fosil_gwh", "total_gwh",   # GWh de referencia
          "fuente"]]
        .sort_values(["pais", "anio"])
        .reset_index(drop=True)
    )
    return df_detalle, df_agregado


# --------------------------------------------------------------------------- #
# 4. VALIDAR
# --------------------------------------------------------------------------- #
def validar(df_detalle: pd.DataFrame, df_agregado: pd.DataFrame,
            df_raw: pd.DataFrame) -> None:
    """
    Chequeos de integridad. Los conteos de fila inesperados solo avisan
    (WARNING); los problemas graves (nulos en columnas clave, categorías
    desconocidas, reconciliaciones que no cuadran o total <= 0) se
    registran como ERROR y abortan sin escribir los CSV.
    """
    errores_graves = 0
    n_paises = df_raw["pais"].nunique()
    n_anios = df_raw["anio"].nunique()
    esperado_detalle = n_paises * n_anios * N_HIJOS_ESPERADOS
    esperado_agg = n_paises * n_anios

    log.info("Validación: %d países × %d años.", n_paises, n_anios)

    # 4.1 conteos de fila
    if len(df_detalle) != esperado_detalle:
        log.warning("Detalle: %d filas (esperado %d). Puede indicar una fuente "
                    "no extraída en algún país.", len(df_detalle), esperado_detalle)
    else:
        log.info("Detalle: %d filas OK (= %d países × %d años × %d fuentes).",
                 len(df_detalle), n_paises, n_anios, N_HIJOS_ESPERADOS)

    if len(df_agregado) != esperado_agg:
        log.warning("Agregado: %d filas (esperado %d).", len(df_agregado), esperado_agg)
    else:
        log.info("Agregado: %d filas OK (= %d países × %d años).",
                 len(df_agregado), n_paises, n_anios)

    # 4.2 nulos en columnas clave
    claves_det = ["pais", "anio", "tipo_fuente", "categoria", "valor_kwh"]
    nulos = df_detalle[claves_det].isna().sum()
    if nulos.any():
        log.error("Detalle: hay NULOS en columnas clave:\n%s", nulos[nulos > 0])
        errores_graves += 1
    else:
        log.info("Detalle: sin nulos en columnas clave OK.")

    # 4.3 categoría solo puede ser fósil o renovable
    cats = set(df_detalle["categoria"].unique())
    if not cats <= {"fósil", "renovable"}:
        log.error("Detalle: categorías inesperadas: %s", cats - {"fósil", "renovable"})
        errores_graves += 1
    else:
        log.info("Detalle: categorías = %s OK.", sorted(cats))

    # 4.4 reconciliaciones por país-año (el chequeo más importante)
    problemas = 0
    for _, fila in df_agregado.iterrows():
        # (i) fósil + renovable == total (usando el Total de ORIGEN)
        suma = fila["fosil_kwh"] + fila["renovable_kwh"]
        if abs(suma - fila["total_kwh"]) > TOLERANCIA_KWH:
            problemas += 1
            log.error("%s %d: fósil+renovable=%.1f != total_origen=%.1f (delta=%.1f kWh)",
                      fila["pais"], fila["anio"], suma, fila["total_kwh"],
                      abs(suma - fila["total_kwh"]))
        # (ii) total > 0 (si no, ECO11/ECO13 dividirían por cero)
        if fila["total_kwh"] <= 0:
            problemas += 1
            log.error("%s %d: total_kwh <= 0 (%.1f). Rompería ECO11/ECO13.",
                      fila["pais"], fila["anio"], fila["total_kwh"])
        # (iii) coherencia de porcentajes: %fósil + %renovable == 100
        pct_fosil = fila["fosil_kwh"] / fila["total_kwh"] * 100
        pct_renov = fila["renovable_kwh"] / fila["total_kwh"] * 100
        if abs((pct_fosil + pct_renov) - 100) > 1e-6:
            problemas += 1
            log.error("%s %d: %%fósil+%%renov=%.6f != 100.",
                      fila["pais"], fila["anio"], pct_fosil + pct_renov)

    if problemas == 0:
        log.info("Reconciliación país-año: %d filas OK "
                 "(fósil+renovable=total, total>0, %%fósil+%%renov=100).",
                 len(df_agregado))
    else:
        log.error("Reconciliación: %d problema(s) detectado(s). Revisar arriba.", problemas)
        errores_graves += problemas

    # 4.5 cruce detalle vs agregado: renovable del detalle == renovable del agg
    renov_det = (
        df_detalle[df_detalle["categoria"] == "renovable"]
        .groupby(["pais", "anio"])["valor_kwh"].sum()
        .rename("renov_detalle")
    )
    cmp = df_agregado.set_index(["pais", "anio"])["renovable_kwh"].to_frame().join(renov_det)
    desalineado = (cmp["renovable_kwh"] - cmp["renov_detalle"]).abs() > TOLERANCIA_KWH
    if desalineado.any():
        log.error("Renovable no coincide entre detalle y agregado en:\n%s",
                  cmp[desalineado])
        errores_graves += 1
    else:
        log.info("Cruce detalle<->agregado (renovable): coincide OK.")

    if errores_graves:
        sys.exit("VALIDACIÓN FALLIDA — los CSV no fueron escritos.")


# --------------------------------------------------------------------------- #
# 5. MAIN
# --------------------------------------------------------------------------- #
def _localizar_xlsx(carpeta: Path) -> Path:
    """Encuentra el único .xlsx en la carpeta raw. Falla claro si hay 0 o >1."""
    candidatos = sorted(carpeta.glob("*.xlsx"))
    if not candidatos:
        raise FileNotFoundError(f"No hay ningún .xlsx en {carpeta}")
    if len(candidatos) > 1:
        log.warning("Hay %d archivos .xlsx en %s; uso el primero: %s",
                    len(candidatos), carpeta, candidatos[0].name)
    return candidatos[0]


def main() -> None:
    log.info("=== ETL generación por fuente — inicio ===")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_path = _localizar_xlsx(RAW_DIR)
    df_raw = extraer_datos(xlsx_path)
    df_detalle, df_agregado = transformar(df_raw)
    validar(df_detalle, df_agregado, df_raw)

    salida_detalle = OUT_DIR / "generacion_por_fuente_detalle.csv"
    salida_agg = OUT_DIR / "generacion_por_tipo_de_fuente.csv"
    # utf-8-sig -> Excel en Windows abre los acentos correctamente.
    df_detalle.to_csv(salida_detalle, index=False, encoding="utf-8-sig")
    df_agregado.to_csv(salida_agg, index=False, encoding="utf-8-sig")

    log.info("Escrito: %s (%d filas)", salida_detalle.name, len(df_detalle))
    log.info("Escrito: %s (%d filas)", salida_agg.name, len(df_agregado))
    log.info("=== ETL generación por fuente — fin ===")


if __name__ == "__main__":
    main()
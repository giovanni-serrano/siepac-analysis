"""
generar_matriz_indicadores.py — Indicadores de la dimensión económica
====================================================
Etapa del pipeline : indicadores
Entradas           : data/processed/matriz_consolidada_wide.csv y
                     data/processed/matriz_consolidada_tidy.csv
Salidas            : data/processed/indicadores_ECO_SIEPAC.xlsx (fórmulas
                     auditables) y
                     data/processed/indicadores_ECO_valores.csv (valores
                     planos para los visualizadores)
Alimenta           : ECO1, ECO2, ECO3, ECO6, ECO11, ECO13, ECO14, ECO15
Fuente de datos    : matriz consolidada (OLADE, CEPAL, Banco Mundial)

Uso:  python src/generar_matriz_indicadores.py   (ejecutar desde la raíz)

Notas metodológicas:
  - En el Excel, las celdas de resultado contienen FÓRMULAS que
    referencian la hoja Datos_Base (no valores pegados), para que cada
    resultado sea auditable haciendo clic en la celda.
  - El CSV de valores existe porque un .xlsx escrito por openpyxl guarda
    las fórmulas SIN resultado cacheado (solo Microsoft Excel calcula y
    cachea al guardar): leer las hojas ECO con pandas devolvería NaN.
  - Cada entrada de INDICADORES define la fórmula de Excel y su cálculo
    pandas equivalente lado a lado; deben mantenerse en paridad.

Estructura del libro:
  1. Metodologia    — índice, fórmula, unidad y fuente por indicador
  2. Datos_Base     — matriz consolidada wide (datos de entrada)
  3. ECO1 ... ECO15 — una hoja por indicador (país × año, con fórmulas)

Autor: Luis Giovanni Serrano Bello — Tesis SIEPAC, UNI Nicaragua
"""

import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config_siepac import PAISES_SIEPAC as PAISES, ANIOS_ANALISIS as ANIOS, DIR_PROCESSED

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-7s | %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger(Path(__file__).stem)

RUTA_WIDE = DIR_PROCESSED / "matriz_consolidada_wide.csv"
RUTA_TIDY = DIR_PROCESSED / "matriz_consolidada_tidy.csv"
RUTA_SALIDA = DIR_PROCESSED / "indicadores_ECO_SIEPAC.xlsx"
RUTA_VALORES = DIR_PROCESSED / "indicadores_ECO_valores.csv"

# ----------------------------- estilos sobrios -----------------------------
FUENTE_BASE = Font(name="Arial", size=10)
FUENTE_TITULO = Font(name="Arial", size=12, bold=True, color="1F3864")
FUENTE_SUB = Font(name="Arial", size=9, italic=True, color="595959")
FUENTE_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill("solid", start_color="44546A")   # azul grisaceo sobrio
FILL_PROMEDIO = PatternFill("solid", start_color="EDEDED")
FILL_IMPUTADO = PatternFill("solid", start_color="FFF2CC") # amarillo palido
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)
CENTRO = Alignment(horizontal="center", vertical="center")
IZQ = Alignment(horizontal="left", vertical="center")

# --------------------- definicion de los 8 indicadores ---------------------
# Columnas de Datos_Base (A=pais, B=anio, luego variables en orden fijo):
COLS_BASE = [
    "pais", "anio",
    "consumo_final_total_kwh",   # C
    "consumo_industrial_kwh",    # D
    "exportaciones_kwh",         # E
    "gen_biomasa_kwh",           # F
    "gen_eolica_kwh",            # G
    "gen_fosil_kwh",             # H
    "gen_geotermia_kwh",         # I
    "gen_hidro_kwh",             # J
    "gen_renovable_kwh",         # K
    "gen_solar_kwh",             # L
    "gen_total_kwh",             # M
    "importaciones_kwh",         # N
    "pib_usd_const2015",         # O
    "poblacion_habitantes",      # P
    "produccion_bruta_kwh",      # Q
    "tarifa_usd_mwh",            # R
    "vai_pct_pib",               # S
    "vai_usd_const2015",         # T
    "tarifa_fuente_dato",        # U (real / imputado_CAGR)
]

# Cada indicador: titulo, unidad, plantilla de formula ({r} = fila en
# Datos_Base para ese pais-anio), su calculo pandas EQUIVALENTE (mismo
# resultado que la formula, sobre la matriz wide), formato y descripcion.
# formula y calculo deben mantenerse en paridad: si cambia una, cambia la otra.
#
# formula_agregado: plantilla de la fila "Agregado regional (razon de
# sumas)" — cada {X} se expande a SUM(...) de la columna X de Datos_Base
# sobre los 6 paises del anio. Es el agregado Σnum/Σden (pondera cada
# pais por su denominador), en paridad con viz_comun.agregados_eco.
# ECO14 no la tiene: sin energia regulada vendida (MWh) por pais no hay
# denominador con el que ponderar la tarifa.
INDICADORES = {
    "ECO1": dict(
        titulo="Uso de energía per cápita",
        unidad="kWh/habitante",
        formula="=Datos_Base!C{r}/Datos_Base!P{r}",
        formula_agregado="={C}/{P}",
        calculo=lambda d: d["consumo_final_total_kwh"] / d["poblacion_habitantes"],
        descripcion="Consumo Final Total (kWh) ÷ Población Total (habitantes)",
        fuentes="SIELAC-OLADE (consumo); CEPAL-CELADE (población)",
        num_fmt="#,##0",
    ),
    "ECO2": dict(
        titulo="Uso de energía por unidad de PIB",
        unidad="kWh/USD const. 2015",
        formula="=Datos_Base!C{r}/Datos_Base!O{r}",
        formula_agregado="={C}/{O}",
        calculo=lambda d: d["consumo_final_total_kwh"] / d["pib_usd_const2015"],
        descripcion="Consumo Final Total (kWh) ÷ PIB Real (USD constantes 2015)",
        fuentes="SIELAC-OLADE (consumo); Banco Mundial NY.GDP.MKTP.KD (PIB)",
        num_fmt="0.0000",
    ),
    "ECO3": dict(
        titulo="Eficiencia de la conversión y distribución de energía",
        unidad="%",
        formula="=Datos_Base!C{r}/Datos_Base!Q{r}*100",
        formula_agregado="={C}/{Q}*100",
        calculo=lambda d: d["consumo_final_total_kwh"]
                          / d["produccion_bruta_kwh"] * 100,
        descripcion="(Consumo Final Total ÷ Producción Bruta Total) × 100. "
                     "Aproxima la eficiencia del sistema eléctrico desde "
                     "generación hasta consumo final; no representa la cadena "
                     "energética primaria completa.",
        fuentes="SIELAC-OLADE (ambas variables)",
        num_fmt='0.0"%"',
    ),
    "ECO6": dict(
        titulo="Intensidades energéticas de la industria",
        unidad="kWh/USD const. 2015",
        formula="=Datos_Base!D{r}/Datos_Base!T{r}",
        formula_agregado="={D}/{T}",
        calculo=lambda d: d["consumo_industrial_kwh"] / d["vai_usd_const2015"],
        descripcion="Consumo Final Industrial (kWh) ÷ Valor Agregado Industrial "
                     "(USD constantes 2015). El VAI en USD ya fue calculado en el "
                     "ETL como VAI%% × PIB real.",
        fuentes="SIELAC-OLADE (consumo industrial); CEPALSTAT ODS 9.2.1 × "
                "Banco Mundial (VAI)",
        num_fmt="0.0000",
    ),
    "ECO11": dict(
        titulo="Porcentaje de combustibles fósiles en la electricidad",
        unidad="%",
        formula="=Datos_Base!H{r}/Datos_Base!M{r}*100",
        formula_agregado="={H}/{M}*100",
        calculo=lambda d: d["gen_fosil_kwh"] / d["gen_total_kwh"] * 100,
        descripcion="(Generación Térmica Fósil ÷ Generación Total) × 100",
        fuentes="SIELAC-OLADE (generación por tipo de fuente)",
        num_fmt='0.0"%"',
    ),
    "ECO13": dict(
        titulo="Porcentaje de energías renovables en la electricidad",
        unidad="%",
        formula="=(Datos_Base!J{r}+Datos_Base!I{r}+Datos_Base!G{r}"
                "+Datos_Base!L{r}+Datos_Base!F{r})/Datos_Base!M{r}*100",
        formula_agregado="=({J}+{I}+{G}+{L}+{F})/{M}*100",
        calculo=lambda d: (d["gen_hidro_kwh"] + d["gen_geotermia_kwh"]
                           + d["gen_eolica_kwh"] + d["gen_solar_kwh"]
                           + d["gen_biomasa_kwh"]) / d["gen_total_kwh"] * 100,
        descripcion="(Hidro + Geotermia + Eólica + Solar + Biomasa) ÷ "
                     "Generación Total × 100. La fórmula suma las cinco fuentes "
                     "renovables explícitamente (no usa la columna agregada) "
                     "para que el cálculo sea auditable componente a componente.",
        fuentes="SIELAC-OLADE (generación por tipo de fuente)",
        num_fmt='0.0"%"',
    ),
    "ECO14": dict(
        titulo="Precios de la energía de uso final por sector",
        unidad="USD corrientes/MWh",
        formula="=Datos_Base!R{r}",
        calculo=lambda d: d["tarifa_usd_mwh"],
        descripcion="Ingresos por energía regulada vendida (USD) ÷ energía "
                     "regulada consumida (MWh), calculado en el ETL. En USD "
                     "corrientes del año bajo análisis (no constantes). Las "
                     "celdas sombreadas en amarillo son valores imputados vía "
                     "CAGR de la serie histórica 2015+ (ver Datos_Base, "
                     "columna tarifa_fuente_dato).",
        fuentes="CEPAL-SIECA (serie histórica); imputación CAGR para años "
                "faltantes documentada en el ETL",
        num_fmt="0.00",
    ),
    "ECO15": dict(
        titulo="Dependencia de las importaciones netas de energía",
        unidad="%",
        formula="=(Datos_Base!N{r}-Datos_Base!E{r})/"
                "(Datos_Base!Q{r}+Datos_Base!N{r}-Datos_Base!E{r})*100",
        formula_agregado="=({N}-{E})/({Q}+{N}-{E})*100",
        calculo=lambda d: (d["importaciones_kwh"] - d["exportaciones_kwh"])
                          / (d["produccion_bruta_kwh"]
                             + d["importaciones_kwh"]
                             - d["exportaciones_kwh"]) * 100,
        descripcion="((Importaciones − Exportaciones) ÷ (Producción Bruta + "
                     "Importaciones − Exportaciones)) × 100. Positivo = "
                     "importador neto; negativo = exportador neto.",
        fuentes="SIELAC-OLADE (matriz de balance energético; producción bruta)",
        num_fmt='0.0"%"',
    ),
}


def fila_base(pais: str, anio: int) -> int:
    """Fila en Datos_Base para (pais, anio). Datos ordenados pais→anio,
    encabezado en fila 1, datos desde la fila 2."""
    return 2 + PAISES.index(pais) * len(ANIOS) + ANIOS.index(anio)


def calcular_valores(wide: pd.DataFrame) -> pd.DataFrame:
    """Evalúa en pandas el `calculo` de cada indicador sobre la matriz wide.

    Devuelve un DataFrame pais | anio | ECO1..ECO15 | tarifa_fuente_dato,
    con los mismos valores que producirían las fórmulas del Excel. Es la
    salida legible por máquina que consumen los visualizadores.
    """
    valores = wide[["pais", "anio"]].copy()
    for codigo, info in INDICADORES.items():
        valores[codigo] = info["calculo"](wide)
    # Bandera de imputación de la tarifa (la usan ECO14 y los visualizadores).
    valores["tarifa_fuente_dato"] = wide["tarifa_fuente_dato"]
    return valores


def hoja_datos_base(wb: Workbook, wide: pd.DataFrame) -> None:
    ws = wb.create_sheet("Datos_Base")
    ws.append(COLS_BASE)
    for c in range(1, len(COLS_BASE) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font, cel.fill, cel.alignment, cel.border = (
            FUENTE_HEADER, FILL_HEADER, CENTRO, BORDE)

    for _, fila in wide.iterrows():
        ws.append([fila[c] for c in COLS_BASE])

    formatos = {  # formato numerico por columna (letra)
        "C": "#,##0", "D": "#,##0", "E": "#,##0", "F": "#,##0", "G": "#,##0",
        "H": "#,##0", "I": "#,##0", "J": "#,##0", "K": "#,##0", "L": "#,##0",
        "M": "#,##0", "N": "#,##0", "O": "#,##0", "P": "#,##0", "Q": "#,##0",
        "R": "0.00", "S": "0.0", "T": "#,##0",
    }
    for r in range(2, 2 + len(wide)):
        for c in range(1, len(COLS_BASE) + 1):
            cel = ws.cell(row=r, column=c)
            cel.font, cel.border = FUENTE_BASE, BORDE
            letra = get_column_letter(c)
            if letra in formatos:
                cel.number_format = formatos[letra]
            if letra == "B":
                cel.number_format = "0"  # anio sin separador de miles
        # sombrear tarifa imputada
        if ws.cell(row=r, column=21).value == "imputado_CAGR":
            ws.cell(row=r, column=18).fill = FILL_IMPUTADO

    anchos = {"A": 13, "B": 7} | {get_column_letter(c): 16 for c in range(3, 22)}
    for letra, w in anchos.items():
        ws.column_dimensions[letra].width = w
    ws.freeze_panes = "C2"


def hoja_indicador(wb: Workbook, codigo: str, info: dict,
                   imputados: set[tuple[str, int]]) -> None:
    ws = wb.create_sheet(codigo)
    ws["A1"] = f"{codigo} — {info['titulo']} ({info['unidad']})"
    ws["A1"].font = FUENTE_TITULO
    ws["A2"] = f"Fórmula: {info['descripcion'].split('.')[0]}"
    ws["A2"].font = FUENTE_SUB

    # encabezado (fila 3)
    ws.cell(row=3, column=1, value="País")
    for j, anio in enumerate(ANIOS):
        ws.cell(row=3, column=2 + j, value=anio)
    for c in range(1, 2 + len(ANIOS)):
        cel = ws.cell(row=3, column=c)
        cel.font, cel.fill, cel.alignment, cel.border = (
            FUENTE_HEADER, FILL_HEADER, CENTRO, BORDE)
        cel.number_format = "0"

    # cuerpo: una fila por pais, formulas hacia Datos_Base
    for i, pais in enumerate(PAISES):
        fila = 4 + i
        cp = ws.cell(row=fila, column=1, value=pais)
        cp.font, cp.border, cp.alignment = FUENTE_BASE, BORDE, IZQ
        for j, anio in enumerate(ANIOS):
            cel = ws.cell(row=fila, column=2 + j)
            cel.value = info["formula"].format(r=fila_base(pais, anio))
            cel.font, cel.border = FUENTE_BASE, BORDE
            cel.number_format = info["num_fmt"]
            if codigo == "ECO14" and (pais, anio) in imputados:
                cel.fill = FILL_IMPUTADO

    # fila de promedio de paises (formula AVERAGE sobre la propia hoja)
    fila_prom = 4 + len(PAISES)
    cp = ws.cell(row=fila_prom, column=1,
                 value="Promedio de países (media simple)")
    cp.font = Font(name="Arial", size=10, bold=True)
    cp.fill, cp.border = FILL_PROMEDIO, BORDE
    for j in range(len(ANIOS)):
        col = get_column_letter(2 + j)
        cel = ws.cell(row=fila_prom, column=2 + j,
                      value=f"=AVERAGE({col}4:{col}{fila_prom - 1})")
        cel.font = Font(name="Arial", size=10, bold=True)
        cel.fill, cel.border = FILL_PROMEDIO, BORDE
        cel.number_format = info["num_fmt"]

    # fila de agregado regional (razon de sumas Σnum/Σden): formulas SUM
    # hacia Datos_Base, tan auditables como las celdas por pais. En
    # paridad con viz_comun.agregados_eco.
    fila_agr = fila_prom + 1
    if info.get("formula_agregado"):
        ca = ws.cell(row=fila_agr, column=1,
                     value="Agregado regional (razón de sumas)")
        ca.font = Font(name="Arial", size=10, bold=True)
        ca.fill, ca.border = FILL_PROMEDIO, BORDE
        for j, anio in enumerate(ANIOS):
            sumas = {letra: "SUM(" + ",".join(
                         f"Datos_Base!{letra}{fila_base(p, anio)}"
                         for p in PAISES) + ")"
                     for letra in "CDEFGHIJKLMNOPQRST"}
            cel = ws.cell(row=fila_agr, column=2 + j,
                          value=info["formula_agregado"].format(**sumas))
            cel.font = Font(name="Arial", size=10, bold=True)
            cel.fill, cel.border = FILL_PROMEDIO, BORDE
            cel.number_format = info["num_fmt"]
    else:
        nota_agr = ws.cell(row=fila_agr + 1, column=1,
                           value="Sin agregado regional: la fuente no "
                                 "publica la energía regulada vendida "
                                 "(MWh) por país para ponderar la tarifa.")
        nota_agr.font = FUENTE_SUB

    if codigo == "ECO14":
        nota = ws.cell(row=fila_agr + 3, column=1,
                       value="Celdas en amarillo: valores imputados vía CAGR "
                             "(no observados). Detalle en Datos_Base, columna "
                             "tarifa_fuente_dato.")
        nota.font = FUENTE_SUB

    ws.column_dimensions["A"].width = 30
    for j in range(len(ANIOS)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 13
    ws.freeze_panes = "B4"


def hoja_metodologia(wb: Workbook) -> None:
    ws = wb.create_sheet("Metodologia", 0)
    ws["A1"] = ("Indicadores Energéticos de Desarrollo Sostenible (IEDS) — "
                "Dimensión Económica, SIEPAC 2020–2024")
    ws["A1"].font = FUENTE_TITULO
    ws["A2"] = ("Referencia metodológica: OIEA/NU (2005), Indicadores "
                "energéticos del desarrollo sostenible: directrices y "
                "metodologías. Datos base en la hoja Datos_Base; cada hoja "
                "ECO calcula su indicador con fórmulas que referencian esa "
                "hoja. Cada hoja cierra con dos resúmenes: Promedio de "
                "países (media simple, el país típico) y Agregado regional "
                "(razón de sumas Σnum/Σden, el bloque como sistema).")
    ws["A2"].font = FUENTE_SUB

    cab = ["Indicador", "Nombre", "Unidad de salida", "Fórmula", "Fuentes", "Notas"]
    ws.append([])  # fila 3 vacia
    ws.append(cab)  # fila 4
    for c in range(1, len(cab) + 1):
        cel = ws.cell(row=4, column=c)
        cel.font, cel.fill, cel.alignment, cel.border = (
            FUENTE_HEADER, FILL_HEADER, CENTRO, BORDE)

    for codigo, info in INDICADORES.items():
        partes = info["descripcion"].split(". ", 1)
        ws.append([codigo, info["titulo"], info["unidad"],
                   partes[0], info["fuentes"],
                   partes[1] if len(partes) > 1 else ""])
    for r in range(5, 5 + len(INDICADORES)):
        for c in range(1, len(cab) + 1):
            cel = ws.cell(row=r, column=c)
            cel.font, cel.border = FUENTE_BASE, BORDE
            cel.alignment = Alignment(vertical="top", wrap_text=True)

    for letra, w in {"A": 10, "B": 34, "C": 18, "D": 46, "E": 40, "F": 46}.items():
        ws.column_dimensions[letra].width = w
    ws.freeze_panes = "A5"


def main() -> None:
    wide = pd.read_csv(RUTA_WIDE)
    tidy = pd.read_csv(RUTA_TIDY)

    # traer fuente_dato de la tarifa a la wide (columna U de Datos_Base)
    fd = (tidy[tidy["variable"] == "tarifa_usd_mwh"]
          [["pais", "anio", "fuente_dato"]]
          .rename(columns={"fuente_dato": "tarifa_fuente_dato"}))
    wide = wide.merge(fd, on=["pais", "anio"], how="left")
    wide = wide.sort_values(["pais", "anio"]).reset_index(drop=True)

    imputados = {(p, a) for p, a in
                 fd.loc[fd["tarifa_fuente_dato"] == "imputado_CAGR",
                        ["pais", "anio"]].itertuples(index=False)}

    wb = Workbook()
    wb.remove(wb.active)
    hoja_datos_base(wb, wide)
    for codigo, info in INDICADORES.items():
        hoja_indicador(wb, codigo, info, imputados)
    hoja_metodologia(wb)  # se inserta en posicion 0

    wb.save(RUTA_SALIDA)
    log.info("Guardado: %s", RUTA_SALIDA)
    log.info("Hojas: %s", wb.sheetnames)

    # CSV de valores planos (mismos indicadores, calculados en pandas):
    # insumo de los visualizadores, que no pueden leer las fórmulas del xlsx.
    valores = calcular_valores(wide)
    valores.to_csv(RUTA_VALORES, index=False, encoding="utf-8-sig")
    log.info("Guardado: %s (%d filas)", RUTA_VALORES, len(valores))


if __name__ == "__main__":
    main()
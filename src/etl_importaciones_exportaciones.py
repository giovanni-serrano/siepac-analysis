"""
etl_importaciones_exportaciones.py — Intercambios de electricidad
====================================================
Etapa del pipeline : ETL
Entradas           : data/raw/importaciones_exportaciones/*.xlsx (matrices
                     de balance energético, hojas "AAAA - País")
Salidas            : data/processed/importaciones_exportaciones.csv (tidy)
Alimenta           : ECO15
Fuente de datos    : sieLAC-OLADE (Matriz de balance energético)

Uso:  python src/etl_importaciones_exportaciones.py   (ejecutar desde la raíz)

Notas metodológicas:
  - Diseño defensivo: NADA está hardcodeado por posición. Se buscan la
    fila de encabezados, la columna ELECTRICIDAD y las filas de flujo por
    su texto (normalizado sin acentos); una hoja que no cumple la
    estructura se loggea y se salta — nunca se inventan valores.
  - Celda vacía = flujo inexistente en los balances de OLADE, se
    interpreta como 0 GWh y se deja constancia en el log.
  - Los nombres de país se homologan a la grafía canónica del proyecto
    (con tildes) vía NOMBRE_CANONICO.

Autor: Luis Giovanni Serrano Bello — Tesis SIEPAC, UNI Nicaragua
"""

import logging
import re
import sys
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config_siepac import GWH_A_KWH, DIR_RAW, DIR_PROCESSED

# ---------------------------------------------------------------------------
# Configuración de rutas (relativas a la raíz del proyecto, nunca absolutas)
# ---------------------------------------------------------------------------
CARPETA_RAW = DIR_RAW / "importaciones_exportaciones"
CARPETA_PROCESSED = DIR_PROCESSED
ARCHIVO_SALIDA = CARPETA_PROCESSED / "importaciones_exportaciones.csv"

FUENTE = "sieLAC-OLADE (Matriz de balance energético)"

# Nombre normalizado (sin tildes, minúsculas) -> nombre canónico del proyecto.
NOMBRE_CANONICO = {
    "guatemala":   "Guatemala",
    "el salvador": "El Salvador",
    "honduras":    "Honduras",
    "nicaragua":   "Nicaragua",
    "costa rica":  "Costa Rica",
    "panama":      "Panamá",
}

# Países válidos del SIEPAC (normalizados) — cualquier hoja con otro país se
# loggea y salta. Derivado de NOMBRE_CANONICO para no repetir la lista.
PAISES_SIEPAC = set(NOMBRE_CANONICO)

# Patrón del nombre de hoja: "2020 - Costa Rica", "2024 - Panamá", etc.
PATRON_HOJA = re.compile(r"^\s*(\d{4})\s*-\s*(.+?)\s*$")

# Configuración del logging: consola con nivel INFO.
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-7s | %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger(Path(__file__).stem)


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def normalizar(texto) -> str:
    """Convierte a minúsculas y elimina acentos para comparar textos de forma
    robusta ('IMPORTACIÓN' == 'importacion'). Devuelve '' si el valor es None.

    Esto protege contra variaciones de tildes entre hojas u años distintos.
    """
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    # NFD separa letra y tilde; luego descartamos las marcas diacríticas.
    texto = unicodedata.normalize("NFD", texto)
    return "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")


def a_numero(valor):
    """Convierte el contenido de una celda a float.

    Devuelve (numero, es_vacio):
      - Celda vacía (None o cadena en blanco) -> (0.0, True).
        En los balances energéticos de OLADE, una celda en blanco significa
        'flujo inexistente', no 'dato perdido', por eso se interpreta como 0.
        Aun así se marca como vacía para que quede constancia en el log.
      - Valor numérico -> (float, False).
      - Cualquier otra cosa (texto, error de Excel) -> lanza ValueError para
        que la hoja se rechace explícitamente en vez de colar basura.
    """
    if valor is None or (isinstance(valor, str) and valor.strip() == ""):
        return 0.0, True
    if isinstance(valor, (int, float)):
        return float(valor), False
    raise ValueError(f"Valor no numérico: {valor!r}")


# ---------------------------------------------------------------------------
# EXTRACCIÓN
# ---------------------------------------------------------------------------
def extraer_datos(carpeta_raw: Path) -> list[dict]:
    """Recorre todos los .xlsx de la carpeta raw y extrae, de cada hoja
    válida, importación y exportación de ELECTRICIDAD en GWh.

    Devuelve una lista de diccionarios (un registro por hoja procesada).
    Las hojas problemáticas se loggean y se saltan; nunca detienen el resto.
    """
    archivos = sorted(carpeta_raw.glob("*.xlsx"))
    if not archivos:
        log.error("No se encontraron .xlsx en %s", carpeta_raw)
        sys.exit(1)

    registros = []
    for archivo in archivos:
        log.info("Procesando archivo: %s", archivo.name)
        # data_only=True lee los VALORES calculados, no las fórmulas.
        wb = load_workbook(archivo, data_only=True, read_only=True)

        for nombre_hoja in wb.sheetnames:
            try:
                registro = _extraer_hoja(wb[nombre_hoja], nombre_hoja)
                if registro is not None:
                    registro["archivo"] = archivo.name
                    registros.append(registro)
            except Exception as exc:  # noqa: BLE001 — defensivo a propósito
                # Cualquier hoja rota se reporta y se salta. Jamás se
                # inventan valores ni se aborta el lote completo.
                log.error("  [SALTADA] '%s': %s", nombre_hoja, exc)

        wb.close()
    return registros


def _extraer_hoja(ws, nombre_hoja: str) -> dict | None:
    """Extrae los datos de UNA hoja. Devuelve dict o None si se salta.

    Pasos (todos por búsqueda, nada por posición fija):
      1. Parsear año y país del nombre de la hoja.
      2. Buscar la fila de encabezados (la que contiene 'ELECTRICIDAD').
      3. Verificar que la unidad de esa columna sea GWh.
      4. Buscar las filas IMPORTACIÓN y EXPORTACIÓN en la columna A.
      5. Leer y convertir los dos valores.
    """
    # --- 1. Año y país desde el nombre de la hoja -------------------------
    m = PATRON_HOJA.match(nombre_hoja)
    if not m:
        log.warning("  [SALTADA] '%s': el nombre no sigue el patrón "
                    "'AAAA - País'", nombre_hoja)
        return None
    anio = int(m.group(1))
    pais = m.group(2).strip()

    if normalizar(pais) not in PAISES_SIEPAC:
        log.warning("  [SALTADA] '%s': país '%s' no es miembro del SIEPAC",
                    nombre_hoja, pais)
        return None
    if not (1990 <= anio <= 2100):
        log.warning("  [SALTADA] '%s': año %d fuera de rango plausible",
                    nombre_hoja, anio)
        return None

    # --- 2. Localizar fila de encabezados y columna ELECTRICIDAD ----------
    fila_header, col_elec = None, None
    # Se busca solo en las primeras 15 filas: los encabezados siempre están
    # arriba; limitar la búsqueda evita falsos positivos en notas al pie.
    for fila in ws.iter_rows(min_row=1, max_row=15):
        for celda in fila:
            if normalizar(celda.value) == "electricidad":
                fila_header, col_elec = celda.row, celda.column
                break
        if fila_header:
            break
    if fila_header is None:
        raise ValueError("no se encontró la columna 'ELECTRICIDAD' en las "
                         "primeras 15 filas")

    # --- 3. Verificar unidad (fila inmediatamente bajo el encabezado) ------
    unidad = normalizar(ws.cell(fila_header + 1, col_elec).value)
    if unidad != "gwh":
        raise ValueError(f"la unidad de ELECTRICIDAD es '{unidad}', se "
                         f"esperaba 'gwh' — revisar antes de procesar")

    # --- 4. Localizar filas IMPORTACIÓN y EXPORTACIÓN ----------------------
    fila_imp, fila_exp = None, None
    for fila in ws.iter_rows(min_col=1, max_col=1):
        etiqueta = normalizar(fila[0].value)
        # Coincidencia EXACTA tras normalizar: evita confundir con etiquetas
        # compuestas tipo 'importación de crudo' si aparecieran algún día.
        if etiqueta == "importacion":
            fila_imp = fila[0].row
        elif etiqueta == "exportacion":
            fila_exp = fila[0].row
        if fila_imp and fila_exp:
            break
    if fila_imp is None or fila_exp is None:
        raise ValueError(f"no se encontraron las filas de flujo "
                         f"(importación={fila_imp}, exportación={fila_exp})")

    # --- 5. Leer los valores ------------------------------------------------
    imp, imp_vacio = a_numero(ws.cell(fila_imp, col_elec).value)
    exp, exp_vacio = a_numero(ws.cell(fila_exp, col_elec).value)

    if imp_vacio:
        log.warning("  '%s': IMPORTACIÓN de electricidad vacía "
                    "-> interpretada como 0 GWh", nombre_hoja)
    if exp_vacio:
        log.warning("  '%s': EXPORTACIÓN de electricidad vacía "
                    "-> interpretada como 0 GWh", nombre_hoja)
    if imp < 0 or exp < 0:
        # En esta sección del balance los flujos deben ser >= 0
        # (los negativos aparecen solo en la sección de transformación).
        raise ValueError(f"flujo negativo inesperado: imp={imp}, exp={exp}")

    log.info("  OK '%s': imp=%s GWh, exp=%s GWh",
             nombre_hoja, format(imp, ",.2f"), format(exp, ",.2f"))
    return {
        "pais": pais,
        "anio": anio,
        "importaciones_gwh": imp,
        "exportaciones_gwh": exp,
    }


# ---------------------------------------------------------------------------
# TRANSFORMACIÓN
# ---------------------------------------------------------------------------
def transformar(registros: list[dict]) -> pd.DataFrame:
    """Convierte los registros a un DataFrame tidy con las convenciones del
    proyecto: kWh como unidad base (columna GWh de referencia), nombres de
    país homologados a la grafía canónica (con tildes) y orden país/año.
    """
    df = pd.DataFrame(registros)

    # Homologar nombres de país al nombre canónico del proyecto (con tildes),
    # para que el merge con los demás CSVs no falle por 'Panamá' vs 'Panama'.
    df["pais"] = df["pais"].map(lambda p: NOMBRE_CANONICO[normalizar(p)])

    # Conversión a la unidad base del proyecto: GWh -> kWh.
    df["importaciones_kwh"] = df["importaciones_gwh"] * GWH_A_KWH
    df["exportaciones_kwh"] = df["exportaciones_gwh"] * GWH_A_KWH

    df["fuente"] = FUENTE

    columnas = [
        "pais", "anio",
        "importaciones_gwh", "exportaciones_gwh",
        "importaciones_kwh", "exportaciones_kwh",
        "fuente",
    ]
    return (
        df[columnas]
        .sort_values(["pais", "anio"])
        .reset_index(drop=True)
    )


# ---------------------------------------------------------------------------
# VALIDACIÓN
# ---------------------------------------------------------------------------
def validar(df: pd.DataFrame) -> None:
    """Chequeos de integridad. Cualquier fallo detiene el script con error
    explícito: preferimos un fallo ruidoso a un CSV corrupto y silencioso.
    """
    errores = []

    # 1. Sin nulos en ninguna columna.
    nulos = df.isna().sum()
    if nulos.any():
        errores.append(f"Hay valores nulos:\n{nulos[nulos > 0]}")

    # 2. Sin duplicados país-año (indicaría hojas repetidas en el raw).
    duplicados = df.duplicated(subset=["pais", "anio"])
    if duplicados.any():
        errores.append(f"Pares país-año duplicados:\n"
                       f"{df.loc[duplicados, ['pais', 'anio']]}")

    # 3. Panel completo: cada país debe tener los mismos años.
    tabla = df.pivot_table(index="pais", columns="anio",
                           values="importaciones_gwh", aggfunc="count")
    if tabla.isna().any().any():
        errores.append(f"Panel incompleto (país sin algún año):\n{tabla}")

    # 4. Valores no negativos (los flujos de esta sección son >= 0).
    for col in ["importaciones_gwh", "exportaciones_gwh"]:
        if (df[col] < 0).any():
            errores.append(f"Valores negativos en {col}")

    # 5. Coherencia GWh <-> kWh (detecta errores de conversión).
    if not (df["importaciones_kwh"] == df["importaciones_gwh"] * GWH_A_KWH).all():
        errores.append("Inconsistencia en la conversión GWh->kWh (imp)")
    if not (df["exportaciones_kwh"] == df["exportaciones_gwh"] * GWH_A_KWH).all():
        errores.append("Inconsistencia en la conversión GWh->kWh (exp)")

    if errores:
        for e in errores:
            log.error(e)
        sys.exit("VALIDACIÓN FALLIDA — el CSV no fue escrito.")

    log.info("Validación OK: %d filas, %d países, años %d–%d, sin nulos.",
             len(df), df["pais"].nunique(), df["anio"].min(), df["anio"].max())


# ---------------------------------------------------------------------------
# ORQUESTADOR
# ---------------------------------------------------------------------------
def main() -> None:
    log.info("=== ETL importaciones_exportaciones (ECO15) ===")

    registros = extraer_datos(CARPETA_RAW)
    log.info("Hojas procesadas con éxito: %d", len(registros))
    if not registros:
        sys.exit("Ninguna hoja pudo procesarse — revisar los logs de arriba.")

    df = transformar(registros)
    validar(df)

    CARPETA_PROCESSED.mkdir(parents=True, exist_ok=True)
    df.to_csv(ARCHIVO_SALIDA, index=False, encoding="utf-8-sig")
    log.info("CSV escrito en: %s", ARCHIVO_SALIDA)


if __name__ == "__main__":
    main()
"""
procesar_dimensiones.py — Indicadores de las dimensiones ambiental y social
====================================================
Etapa del pipeline : indicadores
Entradas           : data/raw_equipo/ENVs.xlsx y data/raw_equipo/SOCs.xlsx
                     (entregables del equipo, formatos heterogéneos)
Salidas            : data/processed/indicadores_ENV_SIEPAC.xlsx y
                     data/processed/indicadores_SOC_SIEPAC.xlsx
Alimenta           : ENV1, ENV2, ENV3, ENV6, SOC1, SOC2, SOC3
Fuente de datos    : equipo de tesis (dimensiones ambiental y social)

Uso:  python src/procesar_dimensiones.py   (ejecutar desde la raíz)

Convierte los Excel del equipo a dos libros estandarizados con la MISMA
estructura del libro de la dimensión económica:
  - Hoja Metodologia (indicador, serie, unidad, fórmula, notas)
  - Una hoja por serie: fila 1 título, fila 2 fórmula, fila 3 encabezado
    (País + 2020..2024), filas 4-9 países, fila 10 Promedio regional
  - ENV6 conserva su formato especial (2 series por país), es ilustrativo

Notas metodológicas (decisiones de limpieza documentadas):
  - SOC1 viene con años descendentes y países en columnas -> se
    transpone; la columna TOTAL (suma entre países) se descarta.
  - SOC2: El Salvador 2020 = 0 en la fuente equivale a SIN DATO -> se
    deja vacío. Guatemala presenta valores ~1000× menores que el resto
    (posible error de unidades en los ingresos) -> se plasma tal cual
    y se marca en Metodologia para verificación del equipo.
  - ENV1 intensidad viene como fórmula de Excel -> se lee el valor en
    caché y, si faltara, se recalcula con (A×10⁶)/PIB.

Autor: Luis Giovanni Serrano Bello — Tesis SIEPAC, UNI Nicaragua
"""

import logging
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config_siepac import (PAISES_SIEPAC as PAISES, ANIOS_ANALISIS as ANIOS,
                            RAIZ_PROYECTO, DIR_PROCESSED)

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-7s | %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger(Path(__file__).stem)

RUTA_ENV_RAW = RAIZ_PROYECTO / "data" / "raw_equipo" / "ENVs.xlsx"
RUTA_SOC_RAW = RAIZ_PROYECTO / "data" / "raw_equipo" / "SOCs.xlsx"
DIR_OUT = DIR_PROCESSED

# ------------------------------ estilos -----------------------------------
GRIS = "44546A"
F_TIT = Font(name="Arial", size=12, bold=True, color="1F3864")
F_SUB = Font(name="Arial", size=9, italic=True, color="595959")
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_TXT = Font(name="Arial", size=10)
F_NEG = Font(name="Arial", size=10, bold=True)
FILL_H = PatternFill("solid", start_color=GRIS)
FILL_P = PatternFill("solid", start_color="EDEDED")
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)


# ---------------------------------------------------------------------------
# LECTURA DIMENSION AMBIENTAL
# ---------------------------------------------------------------------------

def leer_env(ruta: Path) -> dict[str, pd.DataFrame]:
    """Devuelve {clave_serie: DataFrame pais x anio} + ENV6 especial."""
    wb = openpyxl.load_workbook(ruta, data_only=True)   # valores, no formulas
    series = {}

    def tabla(hoja, col_valor, recalcular=None):
        ws = wb[hoja]
        filas = list(ws.iter_rows(min_row=2, values_only=True))
        reg = {}
        for f in filas:
            pais, anio = f[0], int(f[1])
            v = f[col_valor]
            if v is None and recalcular:
                v = recalcular(f)
            reg[(pais, anio)] = float(v)
        df = pd.DataFrame(
            [[reg[(p, a)] for a in ANIOS] for p in PAISES],
            index=PAISES, columns=ANIOS)
        return df

    # ENV1: col F (idx 5) per capita; col G (idx 6) intensidad (formula)
    series["ENV1_PC"] = tabla("ENV1", 5)
    series["ENV1_PIB"] = tabla("ENV1", 6,
        recalcular=lambda f: (f[2] * 1_000_000) / f[4])
    # ENV2: 4 salidas (cols G..J -> idx 6..9)
    series["ENV2_SO2_PC"] = tabla("ENV2", 6)
    series["ENV2_PAR_PC"] = tabla("ENV2", 7)
    series["ENV2_SO2_PIB"] = tabla("ENV2", 8)
    series["ENV2_PAR_PIB"] = tabla("ENV2", 9)
    # ENV3: escala electrica g/kWh (ultima col, idx 10)
    series["ENV3"] = tabla("ENV3", 10)

    # ENV6: pares de filas (Inyeccion Biomasa / Saldo MER) por pais
    ws = wb["ENV6"]
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    env6 = []
    pais_actual = None
    for f in filas:
        if f[0]:
            pais_actual = str(f[0]).strip()
        serie = str(f[1]).strip()
        env6.append([pais_actual, serie] + [float(v) for v in f[2:7]])
    series["ENV6"] = pd.DataFrame(
        env6, columns=["pais", "serie"] + ANIOS)

    # Datos base de la dimension: variables de entrada de ENV1/ENV2/ENV3.
    reg = {}
    for f in wb["ENV1"].iter_rows(min_row=2, values_only=True):
        reg[(f[0], int(f[1]))] = {
            "emisiones_gei_10e3t": float(f[2]),
            "poblacion_miles": float(f[3]),
            "pib_usd_const2015": float(f[4])}
    for f in wb["ENV2"].iter_rows(min_row=2, values_only=True):
        reg[(f[0], int(f[1]))].update({
            "so2_10e3t": float(f[2]), "particulas_10e3t": float(f[3])})
    for f in wb["ENV3"].iter_rows(min_row=2, values_only=True):
        reg[(f[0], int(f[1]))].update({
            "produccion_bruta_gwh": float(f[2]),
            "nox_10e3t": float(f[4]), "co_10e3t": float(f[5])})
    base = pd.DataFrame(
        [{"pais": p, "anio": a, **reg[(p, a)]}
         for p in PAISES for a in ANIOS])
    series["BASE"] = base
    return series


# ---------------------------------------------------------------------------
# LECTURA DIMENSION SOCIAL
# ---------------------------------------------------------------------------

def leer_soc(ruta: Path) -> dict[str, pd.DataFrame]:
    wb = openpyxl.load_workbook(ruta, data_only=True)
    series = {}

    # --- SOC1: anios en filas (descendentes), paises en columnas C..H ---
    ws = wb["SOC 1"]
    orden_cols = ["Costa Rica", "El Salvador", "Guatemala",
                  "Honduras", "Nicaragua", "Panamá"]  # fila 2, cols C..H
    reg = {}
    for f in ws.iter_rows(min_row=3, max_row=7, values_only=True):
        anio = int(f[0])
        for j, pais in enumerate(orden_cols):
            reg[(pais, anio)] = float(f[2 + j])
    series["SOC1"] = pd.DataFrame(
        [[reg[(p, a)] for a in ANIOS] for p in PAISES],
        index=PAISES, columns=ANIOS)

    # --- SOC2: hoja TOTAL, dos bloques (promedio / quintil mas pobre) ---
    ws = wb["SOC 2 TOTAL"]
    mapa = {"NIC": "Nicaragua", "CR": "Costa Rica", "HD": "Honduras",
            "ES": "El Salvador", "GUA": "Guatemala", "PAN": "Panamá"}
    filas = list(ws.iter_rows(values_only=True))
    encabezados = filas[1][1:7]          # NIC CR HD ES GUA PAN

    # Las filas de datos son las que empiezan con un anio; las primeras 5
    # pertenecen al bloque "hogar promedio" y las siguientes 5 al bloque
    # "quintil mas pobre" (posiciones robustas ante filas vacias).
    filas_dato = [f for f in filas
                  if isinstance(f[0], (int, float)) and f[0] in ANIOS]
    assert len(filas_dato) == 10, f"SOC2: se esperaban 10 filas de datos, hay {len(filas_dato)}"

    def bloque(grupo):
        reg = {}
        for f in grupo:
            anio = int(f[0])
            for j, sigla in enumerate(encabezados):
                v = f[1 + j]
                # El Salvador 2020 = 0 en la fuente => SIN DATO, no cero.
                if mapa[sigla] == "El Salvador" and anio == 2020 and v == 0:
                    v = None
                reg[(mapa[sigla], anio)] = (None if v is None else float(v))
        return pd.DataFrame(
            [[reg[(p, a)] for a in ANIOS] for p in PAISES],
            index=PAISES, columns=ANIOS)

    series["SOC2_PROM"] = bloque(filas_dato[:5])
    series["SOC2_POBRE"] = bloque(filas_dato[5:])

    # --- SOC3: tabla tidy oculta en columnas K..Y ---
    # Se filtra por contenido (col K = pais valido, col L = anio) porque
    # la hoja mezcla bloques auxiliares y el encabezado cambia de fila.
    ws = wb["SOC 3"]
    reg_r, reg_u = {}, {}
    n_filas = 0
    for f in ws.iter_rows(values_only=True):
        if f[10] in PAISES and f[11] is not None:
            try:
                anio = int(f[11])
            except (TypeError, ValueError):
                continue
            if anio not in ANIOS:
                continue
            pais = f[10]
            reg_r[(pais, anio)] = float(f[23])  # col X: acceso renov. rural
            reg_u[(pais, anio)] = float(f[24])  # col Y: acceso renov. urbano
            n_filas += 1
    assert n_filas == 30, f"SOC3: se esperaban 30 filas pais-anio, hay {n_filas}"

    # Datos base de la dimension: tasas de electrificacion y % renovable
    # (insumos de SOC1 y SOC3). Los insumos monetarios de SOC2 (cargo e
    # ingresos) permanecen en las hojas por pais de la fuente, en moneda
    # local, por lo que no se consolidan aqui.
    reg_b = {}
    for f in wb["SOC 3"].iter_rows(values_only=True):
        if f[10] in PAISES and f[11] is not None:
            try:
                anio = int(f[11])
            except (TypeError, ValueError):
                continue
            if anio not in ANIOS:
                continue
            reg_b[(f[10], anio)] = {
                "pct_renovable_generacion": float(f[20]),
                "tasa_electrificacion_rural": float(f[21]),
                "tasa_electrificacion_urbana": float(f[22])}
    for clave, reg in [("SOC3_RURAL", reg_r), ("SOC3_URB", reg_u)]:
        series[clave] = pd.DataFrame(
            [[reg[(p, a)] for a in ANIOS] for p in PAISES],
            index=PAISES, columns=ANIOS)

    series["BASE"] = pd.DataFrame(
        [{"pais": p, "anio": a,
          "pct_sin_electricidad": float(series["SOC1"].loc[p, a]),
          **reg_b[(p, a)]}
         for p in PAISES for a in ANIOS])
    return series


# ---------------------------------------------------------------------------
# ESCRITURA EN FORMATO ESTANDAR (identico al libro ECO)
# ---------------------------------------------------------------------------

def hoja_serie(wb, nombre, df, titulo, formula, num_fmt):
    ws = wb.create_sheet(nombre)
    ws["A1"] = titulo
    ws["A1"].font = F_TIT
    ws["A2"] = f"Fórmula: {formula}"
    ws["A2"].font = F_SUB
    ws.cell(row=3, column=1, value="País")
    for j, a in enumerate(ANIOS):
        ws.cell(row=3, column=2 + j, value=a)
    for c in range(1, 7):
        cel = ws.cell(row=3, column=c)
        cel.font, cel.fill, cel.border = F_HDR, FILL_H, BORDE
        cel.alignment = Alignment(horizontal="center")
    for i, pais in enumerate(PAISES):
        ws.cell(row=4 + i, column=1, value=pais).font = F_TXT
        ws.cell(row=4 + i, column=1).border = BORDE
        for j, a in enumerate(ANIOS):
            v = df.loc[pais, a]
            cel = ws.cell(row=4 + i, column=2 + j,
                          value=None if pd.isna(v) else float(v))
            cel.font, cel.border, cel.number_format = F_TXT, BORDE, num_fmt
    fila_p = 4 + len(PAISES)
    ws.cell(row=fila_p, column=1, value="Promedio regional").font = F_NEG
    ws.cell(row=fila_p, column=1).fill = FILL_P
    ws.cell(row=fila_p, column=1).border = BORDE
    for j, a in enumerate(ANIOS):
        cel = ws.cell(row=fila_p, column=2 + j, value=float(df[a].mean()))
        cel.font, cel.fill, cel.border = F_NEG, FILL_P, BORDE
        cel.number_format = num_fmt
    ws.column_dimensions["A"].width = 20
    for c in "BCDEF":
        ws.column_dimensions[c].width = 13
    ws.freeze_panes = "B4"


def hoja_env6(wb, df):
    ws = wb.create_sheet("ENV6")
    ws["A1"] = "ENV6 — Comparativo de inyección de Biomasa vs Saldo MER (GWh)"
    ws["A1"].font = F_TIT
    ws["A2"] = ("Indicador ilustrativo: contrasta la generación con biomasa "
                "de cada país con su saldo neto en el Mercado Eléctrico "
                "Regional. Saldo negativo = importador neto en el MER.")
    ws["A2"].font = F_SUB
    enc = ["País", "Serie"] + ANIOS
    for c, v in enumerate(enc, start=1):
        cel = ws.cell(row=3, column=c, value=v)
        cel.font, cel.fill, cel.border = F_HDR, FILL_H, BORDE
        cel.alignment = Alignment(horizontal="center")
    for i, fila in df.iterrows():
        for c, v in enumerate(fila, start=1):
            cel = ws.cell(row=4 + i, column=c, value=v)
            cel.font, cel.border = F_TXT, BORDE
            if c > 2:
                cel.number_format = "#,##0.0"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    for c in "CDEFG":
        ws.column_dimensions[c].width = 12
    ws.freeze_panes = "C4"


def hoja_metodologia(wb, filas, titulo):
    ws = wb.create_sheet("Metodologia", 0)
    ws["A1"] = titulo
    ws["A1"].font = F_TIT
    enc = ["Hoja", "Indicador / Serie", "Unidad", "Fórmula", "Notas"]
    for c, v in enumerate(enc, start=1):
        cel = ws.cell(row=3, column=c, value=v)
        cel.font, cel.fill, cel.border = F_HDR, FILL_H, BORDE
    for r, fila in enumerate(filas, start=4):
        for c, v in enumerate(fila, start=1):
            cel = ws.cell(row=r, column=c, value=v)
            cel.font, cel.border = F_TXT, BORDE
            cel.alignment = Alignment(vertical="top", wrap_text=True)
    for letra, w in {"A": 14, "B": 38, "C": 18, "D": 44, "E": 46}.items():
        ws.column_dimensions[letra].width = w
    ws.freeze_panes = "A4"


# Catalogo: (clave, titulo, unidad, formula, formato excel, nota)
CAT_ENV = [
    ("ENV1_PC", "ENV1 — Emisiones GEI per cápita",
     "t CO2eq/habitante",
     "Emisiones GEI de centrales eléctricas (10³ t) ÷ Población", "0.0000",
     "Solo emisiones del sector de generación eléctrica."),
    ("ENV1_PIB", "ENV1 — Emisiones GEI por unidad de PIB",
     "kg CO2eq/USD const. 2015",
     "Emisiones GEI (kg) ÷ PIB real (USD constantes 2015)", "0.0000",
     "Valor recuperado de la fórmula original =(A×10⁶)/PIB."),
    ("ENV2_SO2_PC", "ENV2 — SO₂ per cápita", "kg/habitante",
     "Emisiones SO₂ de centrales eléctricas ÷ Población", "0.000", ""),
    ("ENV2_PAR_PC", "ENV2 — Partículas per cápita", "kg/habitante",
     "Emisiones de partículas ÷ Población", "0.0000", ""),
    ("ENV2_SO2_PIB", "ENV2 — SO₂ por unidad de PIB", "g/USD const. 2015",
     "Emisiones SO₂ (g) ÷ PIB real", "0.0000", ""),
    ("ENV2_PAR_PIB", "ENV2 — Partículas por unidad de PIB",
     "g/USD const. 2015", "Emisiones de partículas (g) ÷ PIB real",
     "0.00000", ""),
    ("ENV3", "ENV3 — Emisiones atmosféricas de los sistemas energéticos",
     "g/kWh", "(SO₂ + NOx + CO + PAR) ÷ Producción bruta", "0.000",
     "Escala eléctrica (g/kWh)."),
]
CAT_SOC = [
    ("SOC1", "SOC1 — Población sin acceso a electricidad", "%",
     "100 − Tasa de electrificación total", "0.00",
     "La columna TOTAL de la fuente (suma entre países) se descartó; "
     "el promedio regional se recalcula como media simple."),
    ("SOC2_PROM", "SOC2 — Ingreso destinado a electricidad (hogar promedio)",
     "%", "Cargo anual de electricidad ÷ Ingreso anual del hogar promedio "
     "× 100", "0.00",
     "Regla de salvaguarda: un 0 en la fuente se trata como sin dato. "
     "ADVERTENCIA: Guatemala presenta valores ~1000× menores que el "
     "resto del bloque; posible inconsistencia de unidades en los "
     "ingresos de la fuente. Verificar con el equipo antes de usar."),
    ("SOC2_POBRE", "SOC2 — Ingreso destinado a electricidad (quintil más "
     "pobre)", "%",
     "Cargo anual ÷ Ingreso anual del quintil de menores ingresos × 100",
     "0.00", "Mismas observaciones que SOC2 (hogar promedio)."),
    ("SOC3_RURAL", "SOC3 — Hogares rurales con acceso a energía renovable",
     "%", "Tasa de electrificación rural × % renovable de la generación",
     "0.00", "Proxy elaborado por el equipo: asume mix uniforme de la red."),
    ("SOC3_URB", "SOC3 — Hogares urbanos con acceso a energía renovable",
     "%", "Tasa de electrificación urbana × % renovable de la generación",
     "0.00", "Proxy elaborado por el equipo: asume mix uniforme de la red."),
]


def hoja_base(wb, base_df, titulo, nota):
    """Hoja Datos_Base: variables de entrada de la dimension, en tidy
    (una fila por pais-anio), mismo patron de estilo que el libro ECO."""
    ws = wb.create_sheet("Datos_Base")
    ws["A1"] = titulo
    ws["A1"].font = F_TIT
    ws["A2"] = nota
    ws["A2"].font = F_SUB
    for c, col in enumerate(base_df.columns, start=1):
        cel = ws.cell(row=3, column=c, value=col)
        cel.font, cel.fill, cel.border = F_HDR, FILL_H, BORDE
        cel.alignment = Alignment(horizontal="center")
    for r, fila in enumerate(base_df.itertuples(index=False), start=4):
        for c, v in enumerate(fila, start=1):
            cel = ws.cell(row=r, column=c, value=v)
            cel.font, cel.border = F_TXT, BORDE
            if c > 2:
                cel.number_format = "#,##0.00"
            if c == 2:
                cel.number_format = "0"
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(base_df.columns) + 1):
        letra = ws.cell(row=3, column=c).column_letter
        ws.column_dimensions[letra].width = 22
    ws.freeze_panes = "C4"


def main():
    log.info("Leyendo ENVs.xlsx y SOCs.xlsx ...")
    env = leer_env(RUTA_ENV_RAW)
    soc = leer_soc(RUTA_SOC_RAW)
    DIR_OUT.mkdir(parents=True, exist_ok=True)

    # ------- libro ambiental -------
    wb = Workbook(); wb.remove(wb.active)
    for clave, titulo, unidad, formula, fmt, nota in CAT_ENV:
        hoja_serie(wb, clave, env[clave], f"{titulo} ({unidad})",
                   formula, fmt)
    hoja_env6(wb, env["ENV6"])
    hoja_base(wb, env["BASE"],
        "Datos base — dimensión ambiental (variables de entrada)",
        "Emisiones en 10³ t; población en miles; PIB en USD constantes "
        "2015; producción bruta en GWh. Fuente: ENVs.xlsx del equipo.")
    hoja_metodologia(
        wb,
        [[c, t, u, f, n] for c, t, u, f, _, n in CAT_ENV] +
        [["ENV6", "ENV6 — Inyección de biomasa vs saldo MER", "GWh",
          "Series observadas (sin fórmula)",
          "Indicador ilustrativo; el saldo MER negativo indica importador "
          "neto en el mercado regional."]],
        "Indicadores de la dimensión ambiental — SIEPAC 2020–2024")
    ruta_env = DIR_OUT / "indicadores_ENV_SIEPAC.xlsx"
    wb.save(ruta_env)
    log.info("Guardado: %s", ruta_env)

    # ------- libro social -------
    wb = Workbook(); wb.remove(wb.active)
    for clave, titulo, unidad, formula, fmt, nota in CAT_SOC:
        hoja_serie(wb, clave, soc[clave], f"{titulo} ({unidad})",
                   formula, fmt)
    hoja_base(wb, soc["BASE"],
        "Datos base — dimensión social (variables de entrada)",
        "Todas las variables en %. Los insumos monetarios de SOC2 "
        "(cargo medio e ingresos por grupo) permanecen en las hojas por "
        "país de SOCs.xlsx, en moneda local. Fuente: SOCs.xlsx del equipo.")
    hoja_metodologia(
        wb, [[c, t, u, f, n] for c, t, u, f, _, n in CAT_SOC],
        "Indicadores de la dimensión social — SIEPAC 2020–2024")
    ruta_soc = DIR_OUT / "indicadores_SOC_SIEPAC.xlsx"
    wb.save(ruta_soc)
    log.info("Guardado: %s", ruta_soc)


if __name__ == "__main__":
    main()
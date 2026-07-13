"""
etl_produccion_bruta.py — Producción bruta de electricidad
====================================================
Etapa del pipeline : ETL
Entradas           : data/raw/produccion_bruta/*.xlsx (hoja "1.Producción")
Salidas            : data/processed/produccion_bruta.csv (tidy)
Alimenta           : ECO3, ECO15
Fuente de datos    : SIELAC-OLADE (Series de oferta y demanda)

Uso:  python src/etl_produccion_bruta.py   (ejecutar desde la raíz del proyecto)

Notas metodológicas:
  - Estructura del raw (verificada por inspección): una sola hoja con 5
    bloques, uno por año; cada bloque trae fila título "Producción - AAAA",
    encabezados (23 energéticos), unidades y 6 filas de países.
  - Decisión metodológica (validada con tutor): no existe una fila
    "Producción Bruta Total"; se usa la columna "Electricidad" (GWh),
    que corresponde a la generación eléctrica bruta de cada país.

Autor: Luis Giovanni Serrano Bello — Tesis SIEPAC, UNI Nicaragua
"""

import logging
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

from config_siepac import GWH_A_KWH, PAISES_SIEPAC, DIR_RAW, DIR_PROCESSED

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

RUTA_RAW = DIR_RAW / "produccion_bruta"
RUTA_SALIDA = DIR_PROCESSED / "produccion_bruta.csv"

NOMBRE_HOJA = "1.Producción"

# Columna del archivo que representa la producción bruta eléctrica
COLUMNA_OBJETIVO = "electricidad"   # comparación en minúsculas
UNIDAD_ESPERADA = "gwh"             # si la unidad cambia, el bloque se salta

# Lista blanca: solo estas filas se aceptan como datos.
# Cualquier otra primera celda (títulos, "Fuente:", vacíos) corta el bloque.
PAISES_VALIDOS = set(PAISES_SIEPAC)

# Título de bloque: "Producción - 2020" (acepta guion normal o largo)
PATRON_TITULO = re.compile(r"^Producci[oó]n\s*[-–]\s*(\d{4})$")

FUENTE = "SIELAC-OLADE"

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-7s | %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger(Path(__file__).stem)


# ---------------------------------------------------------------------------
# Extracción
# ---------------------------------------------------------------------------

def extraer_datos(ruta_carpeta: Path) -> list[list]:
    """Lee la hoja completa del Excel y la devuelve como lista de filas.

    Devuelve la matriz cruda (lista de listas) para que la transformación
    trabaje sobre datos en memoria y sea fácil de testear.
    """
    archivos = sorted(ruta_carpeta.glob("*.xlsx"))
    if not archivos:
        raise FileNotFoundError(f"No hay archivos .xlsx en {ruta_carpeta}")
    if len(archivos) > 1:
        log.warning("Hay %d archivos .xlsx; se usará el primero: %s",
                    len(archivos), archivos[0].name)

    ruta = archivos[0]
    log.info("Leyendo: %s", ruta.name)

    # data_only=True -> valores calculados, no fórmulas
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    if NOMBRE_HOJA not in wb.sheetnames:
        raise ValueError(
            f"No existe la hoja '{NOMBRE_HOJA}'. Hojas: {wb.sheetnames}"
        )
    hoja = wb[NOMBRE_HOJA]
    filas = [list(fila) for fila in hoja.iter_rows(values_only=True)]
    wb.close()
    log.info("Hoja leída: %d filas x %d columnas (máx)",
             len(filas), max((len(f) for f in filas), default=0))
    return filas


# ---------------------------------------------------------------------------
# Transformación
# ---------------------------------------------------------------------------

def _celda_texto(valor) -> str:
    """Convierte una celda a texto limpio ('' si es None)."""
    return str(valor).strip() if valor is not None else ""


def _a_numero(valor):
    """Intenta convertir una celda a float. Devuelve None si no se puede."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace(",", "")
    try:
        return float(texto)
    except ValueError:
        return None


def _procesar_bloque(filas: list[list], idx_titulo: int, anio: int) -> list[dict]:
    """Procesa un bloque anual. Devuelve registros tidy o [] si el bloque
    no se puede delimitar con confianza (loggea el motivo y lo salta)."""

    # --- Encabezados (fila siguiente al título) ---
    if idx_titulo + 2 >= len(filas):
        log.warning("Bloque %d: no hay filas suficientes tras el título. SALTADO.", anio)
        return []

    encabezados = [_celda_texto(c).lower() for c in filas[idx_titulo + 1]]
    unidades = [_celda_texto(c).lower() for c in filas[idx_titulo + 2]]

    # Localizar la columna 'Electricidad' por nombre, nunca por posición
    try:
        col = encabezados.index(COLUMNA_OBJETIVO)
    except ValueError:
        log.warning("Bloque %d: no se encontró la columna '%s'. SALTADO.",
                    anio, COLUMNA_OBJETIVO)
        return []

    # Verificar la unidad de esa columna antes de convertir
    unidad = unidades[col] if col < len(unidades) else ""
    if unidad != UNIDAD_ESPERADA:
        log.warning("Bloque %d: unidad inesperada '%s' (se esperaba '%s'). SALTADO.",
                    anio, unidad, UNIDAD_ESPERADA)
        return []

    # --- Filas de datos: desde título+3, mientras sean países SIEPAC ---
    registros = []
    paises_vistos = set()
    for fila in filas[idx_titulo + 3:]:
        pais = _celda_texto(fila[0] if fila else None)
        if pais not in PAISES_VALIDOS:
            break  # fila vacía, "Fuente:", u otro título -> fin del bloque

        if pais in paises_vistos:
            log.warning("Bloque %d: país duplicado '%s'. Fila ignorada.", anio, pais)
            continue
        paises_vistos.add(pais)

        valor_gwh = _a_numero(fila[col] if col < len(fila) else None)
        if valor_gwh is None:
            log.warning("Bloque %d: valor no numérico para %s. Fila ignorada.",
                        anio, pais)
            continue
        if valor_gwh < 0:
            log.warning("Bloque %d: valor negativo (%s) para %s. Fila ignorada.",
                        anio, valor_gwh, pais)
            continue

        registros.append({
            "pais": pais,
            "anio": anio,
            "valor_gwh": valor_gwh,
            "valor_kwh": valor_gwh * GWH_A_KWH,  # regla global: GWh -> kWh
            "fuente": FUENTE,
        })

    faltantes = PAISES_VALIDOS - paises_vistos
    if faltantes:
        log.warning("Bloque %d: países faltantes: %s", anio, sorted(faltantes))

    return registros


def transformar(filas: list[list]) -> pd.DataFrame:
    """Detecta todos los bloques anuales y los convierte a formato tidy."""
    registros = []
    anios_vistos = set()

    for i, fila in enumerate(filas):
        primera = _celda_texto(fila[0] if fila else None)
        m = PATRON_TITULO.match(primera)
        if not m:
            continue

        anio = int(m.group(1))
        if anio in anios_vistos:
            log.warning("Año %d aparece en más de un bloque. Bloque extra SALTADO.", anio)
            continue
        anios_vistos.add(anio)

        log.info("Bloque detectado: año %d (fila %d de la hoja)", anio, i + 1)
        registros.extend(_procesar_bloque(filas, i, anio))

    if not registros:
        raise ValueError("No se extrajo ningún registro. Revisar estructura del archivo.")

    df = pd.DataFrame(registros)
    df = df.sort_values(["anio", "pais"]).reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# Validación
# ---------------------------------------------------------------------------

def validar(df: pd.DataFrame) -> None:
    """Validaciones básicas: nulos, conteo de filas, duplicados, rangos."""
    # 1. Nulos
    nulos = df.isnull().sum()
    if nulos.any():
        raise ValueError(f"Hay valores nulos en el CSV final:\n{nulos[nulos > 0]}")

    # 2. Conteo esperado: países x años detectados
    esperadas = df["anio"].nunique() * len(PAISES_VALIDOS)
    if len(df) != esperadas:
        log.warning("Conteo de filas: %d (se esperaban %d = %d años x %d países)",
                    len(df), esperadas, df["anio"].nunique(), len(PAISES_VALIDOS))
    else:
        log.info("Conteo de filas OK: %d (%d años x %d países)",
                 len(df), df["anio"].nunique(), len(PAISES_VALIDOS))

    # 3. Duplicados país-año
    dup = df.duplicated(subset=["pais", "anio"]).sum()
    if dup:
        raise ValueError(f"Hay {dup} pares país-año duplicados.")

    # 4. Consistencia GWh <-> kWh
    if not (df["valor_kwh"] == df["valor_gwh"] * GWH_A_KWH).all():
        raise ValueError("Inconsistencia en la conversión GWh -> kWh.")

    # 5. Rango físico plausible (generación anual de un país SIEPAC:
    #    entre ~1 y ~30,000 GWh). Valores fuera de rango = alerta.
    fuera = df[(df["valor_gwh"] < 1) | (df["valor_gwh"] > 30_000)]
    if not fuera.empty:
        log.warning("Valores fuera del rango plausible (1-30,000 GWh):\n%s",
                    fuera.to_string(index=False))

    log.info("Validación completada.")


# ---------------------------------------------------------------------------
# Orquestación
# ---------------------------------------------------------------------------

def main() -> None:
    filas = extraer_datos(RUTA_RAW)
    df = transformar(filas)
    validar(df)

    RUTA_SALIDA.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(RUTA_SALIDA, index=False, encoding="utf-8-sig")
    log.info("CSV generado: %s (%d filas)", RUTA_SALIDA, len(df))
    log.info("Vista previa:\n%s\n...\n%s",
             df.head(8).to_string(index=False),
             df.tail(3).to_string(index=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log.error("El ETL falló: %s", exc)
        sys.exit(1)